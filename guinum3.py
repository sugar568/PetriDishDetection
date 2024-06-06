# -*- coding: utf-8 -*-
"""
Created on Wed Sep 20 13:31:26 2023

@author: abcne
"""

# -*- coding: utf-8 -*-
"""
Created on Wed Sep 20 13:24:08 2023

@author: abcne
"""

# Form implementation generated from reading ui file 'process.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!
#
# Subscribe to PyShine Youtube channel for more detail!

# from PIL.ImageQt import ImageQt




from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtGui import QImage
import cv2, imutils
from PyQt5.QtGui import QPainter, QColor
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import xlwt
import xlrd
import xlsxwriter
import openpyxl
from xlutils.copy import copy
from PIL import Image
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(2000, 1000)

        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setText("")
        self.label.setObjectName("label")
        self.horizontalLayout_3.addWidget(self.label)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.verticalSlider = QtWidgets.QSlider(self.centralwidget)
        self.verticalSlider.setOrientation(QtCore.Qt.Vertical)
        self.verticalSlider.setObjectName("verticalSlider")
        self.horizontalLayout.addWidget(self.verticalSlider)
        self.verticalSlider_2 = QtWidgets.QSlider(self.centralwidget)
        self.verticalSlider_2.setOrientation(QtCore.Qt.Vertical)
        self.verticalSlider_2.setObjectName("verticalSlider_2")
        self.horizontalLayout.addWidget(self.verticalSlider_2)
        self.horizontalLayout_3.addLayout(self.horizontalLayout)
        self.gridLayout.addLayout(self.horizontalLayout_3, 0, 0, 1, 2)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)

        self.pushButton_2.setObjectName("pushButton_2")
        self.horizontalLayout_2.addWidget(self.pushButton_2)

        self.arrayupdate = QtWidgets.QPushButton(self.centralwidget)
        self.arrayupdate.setObjectName("arrayupdate")
        self.horizontalLayout_2.addWidget(self.arrayupdate)

        self.writetoexcel = QtWidgets.QPushButton(self.centralwidget)
        self.writetoexcel.setObjectName("writetoexcel")
        self.horizontalLayout_2.addWidget(self.writetoexcel)

        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setObjectName("pushButton")
        self.horizontalLayout_2.addWidget(self.pushButton)
        self.gridLayout.addLayout(self.horizontalLayout_2, 1, 0, 1, 1)
        spacerItem = QtWidgets.QSpacerItem(
            40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem, 1, 1, 1, 1)
        self.gridLayout_2.addLayout(self.gridLayout, 0, 0, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        # colony number displays
        w = QtWidgets.QWidget()
        self.label2 = QtWidgets.QLabel(w)
        self.label2.setText("Blue Colonies")
        self.horizontalLayout_3.addWidget(self.label2)

        self.label3 = QtWidgets.QLabel(w)
        self.label3.setText("Red Colonies")
        self.horizontalLayout_3.addWidget(self.label3)

        self.path = QtWidgets.QLabel(w)
        self.path.setText("enter path")
        self.horizontalLayout_3.addWidget(self.path)

        self.ptext = QtWidgets.QLineEdit()

        self.horizontalLayout_3.addWidget(self.ptext)

        self.retranslateUi(MainWindow)
        # self.verticalSlider.valueChanged['int'].connect(self.brightness_value)
        # self.verticalSlider_2.valueChanged['int'].connect(self.blur_value)
        self.pushButton_2.clicked.connect(self.loadImage)
        self.pushButton.clicked.connect(self.savePhoto)
        self.arrayupdate.clicked.connect(self.updateArrayExcel)
        self.writetoexcel.clicked.connect(self.writeToExcel)

        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        # Added code here
        self.filename = None  # Will hold the image address location
        self.tmp = None  # Will hold the temporary image for display
        self.brightness_value_now = 0  # Updated brightness value
        self.blur_value_now = 0  # Updated blur value
        self.bluecolonies = []
        self.redcolonies = []
        self.redcolonydisplayed = 0
        self.bluecolonydisplayed = 0

    def loadImage(self):
        """ This function will load the user-selected image
            and set it to label using the setPhoto function
        """
        self.filename = QFileDialog.getOpenFileName(filter="Image (*.*)")[0]
        self.image = cv2.imread(self.filename)

        print(self.filename)

        img_gray = cv2.cvtColor(self.image, cv2.COLOR_BGR2GRAY)
        img_rgb = cv2.cvtColor(self.image, cv2.COLOR_BGR2RGB)

        # Use minSize because for not
        # bothering with extra-small
        # dots that would look like STOP signs
        '''
        stop_data = cv2.CascadeClassifier(
            r"C:\Python311\Lib\site-packages\cv2\data\haarcascade_red_dot_two.xml")

        found = stop_data.detectMultiScale(img_gray,
                                           minSize =(20, 20))

        # Don't do anything if there's
        # no sign
        amount_found = len(found)
        '''
        self.setPhoto(self.image)
        '''
        if amount_found != 0:
            print(img_rgb)
            # image2 = Image.open(self.filename)
            # qimage = ImageQt(image2)

            # There may be more than one
            # sign in the image
            for (x, y, width, height) in found:
               # pen = QPen(Qt.red, 3)
               # painter = QPainter()
               # rect=QRect(x,y,width,height)
               # painter.drawRect(rect)
               # pixmap = QtGui.QPixmap.fromImage(qimage)
               # painter.drawPixmap(pixmap)
              # pen = QPen(Qt.red, 3)
              # painter.setPen(pen)

               # painter.drawLine(x,y,width,height)
               cv2.rectangle(img_rgb, (x, y),
                              (x + height, y + width),
                              (0, 255, 0), 5)

        '''

    def writeToExcel(self):
        pathToExcel = self.ptext.text()

    # Load existing workbook
        workbook = openpyxl.load_workbook(pathToExcel)

    # Select the active sheet or create a new one
        sheet = workbook.active

    # Write 'Red Values' and 'Blue Values' to different cells in the first row
        sheet.cell(row=1, column=1, value='Red Values')
        sheet.cell(row=1, column=2, value='Blue Values')

    # Write red colony values
        counterR = 2
        for i in self.redcolonies:
            sheet.cell(row=counterR, column=1, value=str(i))
            counterR += 1

    # Write blue colony values
        counterB = 2
        for i in self.bluecolonies:
            sheet.cell(row=counterB, column=2, value=str(i))
            counterB += 1

    # Save the modified workbook
        workbook.save(pathToExcel)
        
    def updateArrayExcel(self):
        self.bluecolonies.append(self.bluecolonydisplayed);
        self.redcolonies.append(self.redcolonydisplayed)
        print(self.bluecolonies)
        print(self.redcolonies)


    def setPhoto(self, image):
        """ This function will take an image input and resize it
            only for display purposes and convert it to QImage
            to set it at the label.
        """
        self.tmp=image
        image=imutils.resize(image, width = 640)
        frame=cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        image=QImage(
            frame, frame.shape[1], frame.shape[0], frame.strides[0], QImage.Format_RGB888)
        self.label.setPixmap(QtGui.QPixmap.fromImage(image))
        img_gray=cv2.cvtColor(self.tmp, cv2.COLOR_BGR2GRAY)
        img_rgb=cv2.cvtColor(self.tmp, cv2.COLOR_BGR2RGB)


        # red dots
        # Use minSize because for not
        # bothering with extra-small
        # dots that would look like STOP signs
        #C:/Python311/Lib/site-packages/cv2/data/new_red_cascade_5.xml
        #C:\Python311\Lib\site-packages\cv2\data\haarcascade_red_dot_two.xml
        stopred_data=cv2.CascadeClassifier(
            r"C:\Python311\Lib\site-packages\cv2\data\new_red_cascade_5.xml")

        found=stopred_data.detectMultiScale(img_gray,
                                           minSize = (20, 20))
        counter=0
        # Don't do anything if there's
        # no sign
        amount_found=len(found)
        # self.setPhoto(self.image)
        if amount_found != 0:
            print(img_rgb)
            print(found)
            # image2 = Image.open(self.filename)
            # qimage = ImageQt(image2)

            # There may be more than one
            # sign in the image
            for (x, y, width, height) in found:
                print("this")
                counter += 1
            self.label3.setText("Red Colonies: " + str(counter))
            self.redcolonydisplayed=counter










        # bluedots
        # Use minSize because for not
        # bothering with extra-small
        # dots that would look like STOP signs
        stop_data=cv2.CascadeClassifier(
            r"C:\Python311\Lib\site-packages\cv2\data\bluepinkcascade.xml")

        found=stop_data.detectMultiScale(img_gray,
                                           minSize = (20, 20))
        counter=0
        # Don't do anything if there's
        # no sign
        amount_found=len(found)
        # self.setPhoto(self.image)
        if amount_found != 0:
            print(img_rgb)
            print(found)
            # image2 = Image.open(self.filename)
            # qimage = ImageQt(image2)

            # There may be more than one
            # sign in the image
            for (x, y, width, height) in found:
                print("this")
                counter += 1

            self.label2.setText("Blue Colonies: " + str(counter))
            self.bluecolonydisplayed=counter










    def brightness_value(self, value):
        """ This function will take value from the slider
            for the brightness from 0 to 99
        """
        self.brightness_value_now=value
        print('Brightness: ', value)
        self.update()

    def blur_value(self, value):
        """ This function will take value from the slider
            for the blur from 0 to 99
        """
        self.blur_value_now=value
        print('Blur: ', value)
        self.update()

    def changeBrightness(self, img, value):
        """ This function will take an image (img) and the brightness
            value. It will perform the brightness change using OpenCv
            and after split, will merge the img and return it.
        """
        hsv=cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
        h, s, v=cv2.split(hsv)
        lim=255 - value
        v[v > lim]=255
        v[v <= lim] += value
        final_hsv=cv2.merge((h, s, v))
        img=cv2.cvtColor(final_hsv, cv2.COLOR_HSV2BGR)
        return img

    def changeBlur(self, img, value):
        """ This function will take the img image and blur values as inputs.
            After performing blur operation using the OpenCV function, it returns
            the image img.
        """
        kernel_size=(value + 1, value + 1)  # +1 is to avoid 0
        img=cv2.blur(img, kernel_size)
        return img

    def update(self):
        """ This function will update the photo according to the
            current values of blur and brightness and set it to the photo label.
        """
        img=self.changeBrightness(self.image, self.brightness_value_now)
        img=self.changeBlur(img, self.blur_value_now)
        self.setPhoto(img)


    def savePhoto(self):
        """ This function will save the image"""
        # here provide the output file name
        # lets say we want to save the output as a time stamp
        # uncomment the two lines below

        # import time
        # filename = 'Snapshot '+str(time.strftime("%Y-%b-%d at %H.%M.%S %p"))+'.png'

        # Or we can give any name such as output.jpg or output.png as well
        # filename = 'Snapshot.png'

        # Or a much better option is to let the user decide the location and the extension
        # using a file dialog.

        filename=QFileDialog.getSaveFileName(
            filter = "JPG(*.jpg);;PNG(*.png);;TIFF(*.tiff);;BMP(*.bmp)")[0]

        cv2.imwrite(filename, self.tmp)
        print('Image saved as:', self.filename)

    def retranslateUi(self, MainWindow):
        _translate=QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate(
            "MainWindow", "Pyshine photo editor"))
        self.pushButton_2.setText(_translate("MainWindow", "Open"))
        self.arrayupdate.setText(_translate("MainWindow", "use data"))
        self.writetoexcel.setText(_translate("MainWindow", "write to excel"))
        self.pushButton.setText(_translate("MainWindow", "Save"))

# Subscribe to PyShine Youtube channel for more detail!

# WEBSITE: www.pyshine.com

if __name__ == "__main__":
    import sys
    app=QtWidgets.QApplication(sys.argv)
    ex=Ui_MainWindow()
    w=QtWidgets.QMainWindow()
    ex.setupUi(w)
    w.show()
    sys.exit(app.exec_())
